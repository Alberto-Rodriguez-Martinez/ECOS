# -*- coding: utf-8 -*-
"""
Created on Mon Oct  6 11:47:59 2025

@author: Alberto
"""
# acq32.py  (Python 32-bit)
# Requisitos: numpy (evita pandas/pyarrow). Guarda JSON + NPZ.

import json, uuid, csv
from pathlib import Path
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt

def _now_iso():
    try:
        return datetime.now().isoformat()
    except Exception:
        # Fallback sin tz
        return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

def save_experiment_raw_32(
    *,
    specimen,        # dict (estructura exacta de tu espécimen)
    equipment1,      # dict con 'params' incluyendo F_muestreo, Smin, Smax, Slen, WindowLen
    equipment2,      # dict
    protocol,        # dict
    results,         # dict dinámico (T1, T2, C1, C2, Cw, Cl, L, ...)
    Signal_PE,       # np.ndarray (longitud Slen)
    Signal_TT,       # np.ndarray (longitud Slen)
    Signal_Ref,      # np.ndarray (longitud Slen)
    base_dir="data_32"
):
    # Validaciones mínimas
    params = equipment1.get("params", {})
    Smin = int(params["Smin"]); Smax = int(params["Smax"]); Slen = int(params["Slen"])
    if Slen != (Smax - Smin):
        raise ValueError("Slen debe ser Smax - Smin")
    for name, arr in [("Signal_PE", Signal_PE), ("Signal_TT", Signal_TT), ("Signal_Ref", Signal_Ref)]:
        if int(np.asarray(arr).size) != Slen:
            raise ValueError("%s tamaño %d != Slen %d" % (name, np.asarray(arr).size, Slen))

    exp_id = "EXPID-" + uuid.uuid4().hex[:8].upper()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    d = Path(base_dir) / f"{ts}_{exp_id}"
    (d / "signals").mkdir(parents=True, exist_ok=True)

    meta = {
        "schema_version": "raw-32-1.0",
        "experiment": {"id": exp_id, "timestamp_start": _now_iso(), "timestamp_end": _now_iso(), "operator": "Sebas"},
        "specimen": specimen,
        "protocol": protocol,
        "equipment": {
            "device_1_ultrasound": equipment1,
            "device_2_aux": equipment2
        },
        "notes": ""
    }
    # Guardar metadatos/resultado
    (d / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    (d / "results.json").write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    # Guardar señales en un NPZ comprimido (portátil)
    np.savez_compressed(d / "signals" / "signals.npz",
                        Signal_PE=np.asarray(Signal_PE, dtype=np.float64),
                        Signal_TT=np.asarray(Signal_TT, dtype=np.float64),
                        Signal_Ref=np.asarray(Signal_Ref, dtype=np.float64))
    return str(d)

def load_raw32(exp_dir):
    """
    Carga meta.json, results.json y signals.npz desde la carpeta del experimento.
    Devuelve: meta (dict), results (dict), signals (dict de np.ndarray), tiempos (dict de np.ndarray)
    """
    exp_dir = Path(exp_dir)
    meta    = json.loads((exp_dir/"meta.json").read_text(encoding="utf-8"))
    results = json.loads((exp_dir/"results.json").read_text(encoding="utf-8"))
    npz     = np.load(exp_dir/"signals"/"signals.npz")

    # Señales
    signals = {
        "Signal_PE":  npz["Signal_PE"].astype(np.float64),
        "Signal_TT":  npz["Signal_TT"].astype(np.float64),
        "Signal_Ref": npz["Signal_Ref"].astype(np.float64),
    }

    # Eje temporal (si están F_muestreo y Smin en meta)
    params = meta.get("equipment", {}).get("device_1_ultrasound", {}).get("params", {})
    Fs   = float(params.get("F_muestreo", 0.0) or 0.0)
    Smin = int(params.get("Smin", 0) or 0)

    times = {}
    if Fs > 0:
        for name, y in signals.items():
            n0 = Smin
            N  = y.size
            sample_idx = np.arange(n0, n0 + N, dtype=np.int64)
            times[name] = sample_idx / Fs
    else:
        # Si no hay Fs en meta, devolvemos None y trabajas con índices de muestra
        for name, y in signals.items():
            times[name] = None

    return meta, results, signals, times

def plot_signals(signals, times=None, title="Señales"):
    """Grafica las tres señales si tienes matplotlib instalado."""
    plt.figure()
    for name, y in signals.items():
        x = times.get(name) if (times and times.get(name) is not None) else np.arange(y.size)
        plt.plot(x, y, label=name)  # no forzamos colores
    plt.grid(True)
    plt.xlabel("t [s]" if (times and list(times.values())[0] is not None) else "sample")
    plt.ylabel("amplitud [V]")
    plt.title(title)
    plt.legend()
    plt.tight_layout()
    plt.show()
    
def plot_signals_stacked(signals, times=None, title="Señales (3 subplots)"):
    """
    Dibuja Signal_PE, Signal_TT y Signal_Ref en 3 subplots verticales.
    - signals: dict con np.ndarray por señal.
    - times:   dict con eje temporal por señal (o None para usar índice de muestra).
    """
    import numpy as np
    import matplotlib.pyplot as plt

    # Orden preferente; si falta alguna clave, omitimos; si hay extras, las añadimos al final.
    prefer = ["Signal_PE", "Signal_TT", "Signal_Ref"]
    names = [k for k in prefer if k in signals] + [k for k in signals.keys() if k not in prefer]

    fig, axs = plt.subplots(len(names), 1, sharex=True, figsize=(10, 7))
    if len(names) == 1:
        axs = [axs]

    # ¿Tenemos tiempo en segundos?
    has_time = bool(times) and any(times.get(n) is not None for n in names)

    for ax, name in zip(axs, names):
        y = np.asarray(signals[name], dtype=float)
        if times and times.get(name) is not None:
            x = times[name]
            ax.plot(x, y)
        else:
            x = np.arange(y.size)
            ax.plot(x, y)
        ax.set_ylabel(name)     # etiqueta cada subplot con el nombre de la señal
        ax.grid(True)

    axs[-1].set_xlabel("t [s]" if has_time else "sample")
    fig.suptitle(title)
    fig.tight_layout()  # ajusta espacios para que no se solapen títulos/etiquetas
    plt.show()

def quick_stats(y, Fs=None):
    """Estadísticos rápidos de una señal."""
    y = np.asarray(y, dtype=np.float64)
    stats = {
        "N": int(y.size),
        "min": float(np.min(y)),
        "max": float(np.max(y)),
        "rms": float(np.sqrt(np.mean(y**2))),
        "mean": float(np.mean(y)),
        "std": float(np.std(y, ddof=1)) if y.size > 1 else 0.0,
    }
    # pico en frecuencia (aprox) si se conoce Fs
    if Fs and Fs > 0 and y.size > 1:
        Y = np.fft.rfft(y)
        f = np.fft.rfftfreq(y.size, d=1.0/Fs)
        k = int(np.argmax(np.abs(Y)))
        stats["f_peak_Hz"] = float(f[k])
    return stats

def load_latest(base_dir="data_32"):
    """ carga el último experimento."""
    base = Path(base_dir)
    runs = sorted([p for p in base.glob("*_EXPID-*") if p.is_dir()])
    if not runs:
        raise FileNotFoundError(f"No hay experimentos en {base_dir}")
    return load_raw32(runs[-1])

def export_signal_to_csv(exp_dir, name, y, t=None):
    """ exporta señales a CSV"""
    exp_dir = Path(exp_dir)
    out = exp_dir / "signals" / f"{name}.csv"
    if t is not None:
        data = np.column_stack([t, y])
        np.savetxt(out, data, delimiter=",", header="time_s,value", comments="", fmt="%.9g")
    else:
        np.savetxt(out, y, delimiter=",", header="value", comments="", fmt="%.9g")
    return str(out)


def export_results_catalog_csv(base_dir="data_32",
                               out_csv="catalog_results.csv",
                               delimiter=";",
                               decimal_comma=False):
    """
    Recorre <base_dir> y genera un CSV con una fila por experimento usando
    SOLO meta.json (bloque specimen + experiment) y results.json.
    - delimiter: usa ';' si tu Excel está en configuración española.
    - decimal_comma=True: convierte floats '1234.56' -> '1234,56' (como texto).
    """
    base = Path(base_dir)
    runs = sorted([p for p in base.glob("*_EXPID-*") if p.is_dir()])
    rows = []
    specimen_cols, result_cols = set(), set()

    def fmt(v):
        # Convierte floats a coma decimal si se solicita
        if decimal_comma and isinstance(v, float):
            # 12 cifras sig. para no inflar el texto
            s = f"{v:.12g}".replace(".", ",")
            return s
        return v

    for run in runs:
        meta_path = run / "meta.json"
        res_path  = run / "results.json"
        if not (meta_path.exists() and res_path.exists()):
            continue

        meta    = json.loads(meta_path.read_text(encoding="utf-8"))
        results = json.loads(res_path.read_text(encoding="utf-8"))
        specimen = meta.get("specimen", {})
        expinfo  = meta.get("experiment", {})

        row = {
            "experiment_id": expinfo.get("id", ""),
            "timestamp_start": expinfo.get("timestamp_start", ""),
            "timestamp_end": expinfo.get("timestamp_end", ""),
            "operator": expinfo.get("operator", ""),
        }

        # Aplana specimen.* y res.* para evitar colisiones de nombres
        for k, v in specimen.items():
            key = f"specimen.{k}"
            row[key] = fmt(v)
            specimen_cols.add(key)

        for k, v in results.items():
            key = f"res.{k}"
            # Intenta convertir strings numéricas con coma a float
            if isinstance(v, str) and v.count(",") == 1 and v.replace(",", "").replace(".", "").isdigit():
                try:
                    v = float(v.replace(",", "."))
                except Exception:
                    pass
            row[key] = fmt(v)
            result_cols.add(key)

        rows.append(row)

    # Orden de columnas: meta básicas + specimen.* + res.*
    fieldnames = (
        ["experiment_id", "timestamp_start", "timestamp_end", "operator"] +
        sorted(specimen_cols) +
        sorted(result_cols)
    )

    # Escritura del CSV
    out_csv = Path(out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fieldnames})

    return str(out_csv), len(rows), len(fieldnames)

def _extract_row_for_catalog(exp_dir):
    exp_dir = Path(exp_dir)
    meta    = json.loads((exp_dir/"meta.json").read_text(encoding="utf-8"))
    results = json.loads((exp_dir/"results.json").read_text(encoding="utf-8"))

    specimen = meta.get("specimen", {})
    expinfo  = meta.get("experiment", {})

    row = {
        "experiment_id":   expinfo.get("id", ""),
        "timestamp_start": expinfo.get("timestamp_start", ""),
        "timestamp_end":   expinfo.get("timestamp_end", ""),
        "operator":        expinfo.get("operator", ""),
    }
    # Aplana specimen.* y res.* (evita colisiones de nombres)
    for k, v in specimen.items():
        row[f"specimen.{k}"] = v
    for k, v in results.items():
        # intenta convertir strings con coma a float
        if isinstance(v, str):
            try:
                v = float(v.replace(",", ".")) if ("," in v and v.replace(",", ".").replace(".", "", 1).isdigit()) else v
            except Exception:
                pass
        row[f"res.{k}"] = v
    return row

def append_experiment_to_xlsx(exp_dir, xlsx_path="catalog.xlsx", sheet_name="Catalog"):
    """Añade una fila (un experimento) al Excel. Crea el archivo si no existe."""
    from openpyxl import Workbook, load_workbook

    row = _extract_row_for_catalog(exp_dir)
    xlsx_path = Path(xlsx_path)

    # Abrir o crear libro/hoja
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        headers = [c.value for c in (ws[1] if ws.max_row >= 1 else [])]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        headers = []

    # Unir columnas: mantener las existentes y añadir las nuevas al final
    desired = headers[:] if headers else ["experiment_id","timestamp_start","timestamp_end","operator"]
    for k in row.keys():
        if k not in desired:
            desired.append(k)

    # Escribir/actualizar cabecera (primera fila)
    if headers != desired:
        for j, h in enumerate(desired, start=1):
            ws.cell(row=1, column=j, value=h)

    # Escribir fila nueva en el orden de 'desired'
    next_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    for j, h in enumerate(desired, start=1):
        ws.cell(row=next_row, column=j, value=row.get(h, ""))

    wb.save(xlsx_path)
    return str(xlsx_path), next_row-1, len(desired)